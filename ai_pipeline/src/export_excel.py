import sys
import json
import os

def main():
    try:
        # Read JSON from stdin
        input_data = sys.stdin.read()
        transactions = json.loads(input_data)
        
        # Get paths from arguments
        if len(sys.argv) < 3:
            print("Usage: python export_excel.py <template_path> <output_path>")
            sys.exit(1)
            
        template_path = sys.argv[1]
        output_path = sys.argv[2]
        
        from openpyxl import load_workbook
        
        wb = load_workbook(template_path)
        if 'Transactions' not in wb.sheetnames:
            raise ValueError("Sheet 'Transactions' not found in template.")
        
        ws = wb['Transactions']
        start_row = ws.max_row + 1
        
        for row_idx, t in enumerate(transactions, start=start_row):
            ws.cell(row=row_idx, column=1, value=t.get('sourceFile', ''))
            ws.cell(row=row_idx, column=2, value=t.get('date', ''))
            ws.cell(row=row_idx, column=3, value=t.get('description', ''))
            
            def parse_amt(val):
                try:
                    if not val: return None
                    return float(str(val).replace(',', '').replace('$', '').replace('Rs', '').strip())
                except:
                    return val
                    
            ws.cell(row=row_idx, column=4, value=parse_amt(t.get('debit', '')))
            ws.cell(row=row_idx, column=5, value=parse_amt(t.get('credit', '')))
            ws.cell(row=row_idx, column=6, value=parse_amt(t.get('balance', '')))
            ws.cell(row=row_idx, column=7, value='')
            
        wb.save(output_path)
        print("SUCCESS")
        
    except Exception as e:
        print(f"ERROR: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
