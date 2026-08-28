import pandas as pd
from openpyxl import load_workbook

def write_to_excel(transactions, template_path, columns):
    wb = load_workbook(template_path)
    if 'Transactions' not in wb.sheetnames:
        raise ValueError("Sheet 'Transactions' not found in template.")
    
    ws = wb['Transactions']
    
    # Simple append, openpyxl max_row gives the last row with data
    start_row = ws.max_row + 1
    
    for row_idx, t in enumerate(transactions, start=start_row):
        ws.cell(row=row_idx, column=1, value=t.get('sourceFile', '')) # Accounts:
        ws.cell(row=row_idx, column=2, value=t.get('date', ''))
        ws.cell(row=row_idx, column=3, value=t.get('description', ''))
        
        def parse_amt(val):
            try:
                if not val: return None
                return float(str(val).replace(',', '').replace('$', '').replace('Rs', '').strip())
            except:
                return val
                
        ws.cell(row=row_idx, column=4, value=parse_amt(t.get('debit', '')))
        ws.cell(row=row_idx, column=5, value=parse_amt(t.get('credit', '')))
        ws.cell(row=row_idx, column=6, value=parse_amt(t.get('balance', '')))
        ws.cell(row=row_idx, column=7, value='') # Head of Accounts
        
    output_path = template_path.replace('.xlsx', '_Output.xlsx')
    wb.save(output_path)
    print(f"Saved extracted data to {output_path}")
